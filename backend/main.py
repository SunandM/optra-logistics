from fastapi import FastAPI, Depends, HTTPException, Request, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session, joinedload
import uvicorn
from typing import List
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
import os
from openpyxl import Workbook
from openpyxl.styles import Font
import io

from models import get_db, create_tables, User, OTP, UserRequest, ActiveSession
from crud import get_drivers, create_driver, get_vehicles, create_vehicle, get_orders, create_order, get_pending_orders
from schemas import (
    DriverCreate, Driver as DriverSchema, VehicleCreate, Vehicle as VehicleSchema, 
    OrderCreate, Order as OrderSchema, OptimizationRequest, OptimizationResponse,
    UserCreate, UserUpdate, User as UserSchema, UserProfile, LoginRequest, LoginResponse,
    LogoutResponse, PasswordReset, OTPVerify, PasswordResetConfirm, PasswordChange,
    UserRequestCreate, UserRequest as UserRequestSchema, UserRequestUpdate,
    ActiveSessionWithUser
)
from route_optimizer import optimize_routes
from auth_utils import (
    verify_password, get_password_hash, create_access_token, generate_otp, 
    send_otp, ACCESS_TOKEN_EXPIRE_MINUTES
)
from auth_middleware import get_current_user, require_admin_or_higher, require_dispatcher_or_higher, require_superadmin

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Create database tables on startup
    create_tables()
    
    # Seed default superadmin if no users exist
    db = next(get_db())
    try:
        user_count = db.query(User).count()
        if user_count == 0:
            default_admin = User(
                name="Super Admin",
                email="superadmin@routeoptima.com",
                hashed_password=get_password_hash("admin123"),
                role="superadmin",
                is_active=True
            )
            db.add(default_admin)
            db.commit()
            print("Default superadmin created: superadmin@routeoptima.com / admin123")
        
        # Seed dispatcher if doesn't exist
        dispatcher = db.query(User).filter(User.email == "dispatcher@routeoptima.com").first()
        if not dispatcher:
            dispatcher_user = User(
                name="Dispatcher",
                email="dispatcher@routeoptima.com",
                hashed_password=get_password_hash("dispatch123"),
                role="dispatcher",
                is_active=True
            )
            db.add(dispatcher_user)
            db.commit()
            print("Default dispatcher created: dispatcher@routeoptima.com / dispatch123")
        
        # Seed driver if doesn't exist
        driver = db.query(User).filter(User.email == "driver@routeoptima.com").first()
        if not driver:
            driver_user = User(
                name="Driver",
                email="driver@routeoptima.com",
                hashed_password=get_password_hash("driver123"),
                role="driver",
                is_active=True
            )
            db.add(driver_user)
            db.commit()
            print("Default driver created: driver@routeoptima.com / driver123")
    finally:
        db.close()
    
    yield
    # Cleanup on shutdown if needed

app = FastAPI(
    title="Route Optimization API",
    description="A delivery route optimization system similar to MaxOptra",
    version="1.0.0",
    lifespan=lifespan
)

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/health")
async def health_check():
    """Health check endpoint to verify API is running"""
    return JSONResponse(
        status_code=200,
        content={
            "status": "healthy",
            "message": "Route Optimization API is running",
            "version": "1.0.0"
        }
    )

@app.get("/")
async def root():
    """Root endpoint"""
    return {"message": "Route Optimization API", "version": "1.0.0"}

# Authentication endpoints
@app.post("/auth/login", response_model=LoginResponse)
async def login(request: LoginRequest, db: Session = Depends(get_db), http_request: Request = None):
    """User login endpoint"""
    user = db.query(User).filter(User.email == request.email).first()
    if not user or not verify_password(request.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    if not user.is_active:
        raise HTTPException(status_code=401, detail="Account is inactive")
    
    # Update last login
    user.last_login = datetime.utcnow()
    
    # Create active session
    client_ip = http_request.client.host if http_request else "unknown"
    active_session = ActiveSession(
        user_id=user.id,
        login_at=datetime.utcnow(),
        ip_address=client_ip,
        is_active=True
    )
    db.add(active_session)
    db.commit()
    
    # Create access token
    access_token = create_access_token(data={"sub": str(user.id)})
    
    user_profile = UserProfile(
        id=user.id,
        name=user.name,
        email=user.email,
        phone=user.phone,
        role=user.role,
        is_active=user.is_active,
        last_login=user.last_login,
        created_at=user.created_at
    )
    
    return LoginResponse(
        access_token=access_token,
        token_type="bearer",
        user=user_profile
    )

@app.post("/auth/logout", response_model=LogoutResponse)
async def logout(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """User logout endpoint"""
    # Find the current user's most recent active session and mark it as inactive
    session = db.query(ActiveSession).filter(
        ActiveSession.user_id == current_user.id,
        ActiveSession.is_active == True
    ).order_by(ActiveSession.login_at.desc()).first()
    
    if session:
        session.is_active = False
        session.logout_at = datetime.utcnow()
        db.commit()
    
    return LogoutResponse(message="Successfully logged out")

@app.post("/auth/forgot-password")
async def forgot_password(request: PasswordReset, db: Session = Depends(get_db)):
    """Send OTP for password reset"""
    if not request.email and not request.phone:
        raise HTTPException(status_code=400, detail="Email or phone number is required")
    
    # Find user by email or phone
    if request.email:
        user = db.query(User).filter(User.email == request.email).first()
    else:
        user = db.query(User).filter(User.phone == request.phone).first()
    
    if not user:
        # Don't reveal if user exists or not for security
        return {"message": "If the account exists, an OTP has been sent"}
    
    # Generate OTP
    otp_code = generate_otp()
    expires_at = datetime.utcnow() + timedelta(minutes=5)
    
    # Save OTP to database
    otp = OTP(
        user_id=user.id,
        otp_code=otp_code,
        delivery_method=request.delivery_method,
        expires_at=expires_at,
        is_used=False
    )
    db.add(otp)
    db.commit()
    
    # Send OTP
    contact = request.email if request.delivery_method == "email" else request.phone
    send_otp(
        email=request.email if request.delivery_method == "email" else None,
        phone=request.phone if request.delivery_method == "sms" else None,
        otp_code=otp_code,
        delivery_method=request.delivery_method
    )
    
    return {"message": "OTP sent successfully"}

@app.post("/auth/verify-otp")
async def verify_otp(request: OTPVerify, db: Session = Depends(get_db)):
    """Verify OTP code"""
    if not request.email and not request.phone:
        raise HTTPException(status_code=400, detail="Email or phone number is required")
    
    # Find user
    if request.email:
        user = db.query(User).filter(User.email == request.email).first()
    else:
        user = db.query(User).filter(User.phone == request.phone).first()
    
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Find valid OTP
    otp = db.query(OTP).filter(
        OTP.user_id == user.id,
        OTP.otp_code == request.otp_code,
        OTP.is_used == False,
        OTP.expires_at > datetime.utcnow()
    ).first()
    
    if not otp:
        raise HTTPException(status_code=400, detail="Invalid or expired OTP")
    
    # Mark OTP as used
    otp.is_used = True
    db.commit()
    
    return {"message": "OTP verified successfully"}

@app.post("/auth/reset-password")
async def reset_password(request: PasswordResetConfirm, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Reset password using OTP or manual reset by admin"""
    if not request.email and not request.phone:
        raise HTTPException(status_code=400, detail="Email or phone number is required")
    
    # Find user
    if request.email:
        target_user = db.query(User).filter(User.email == request.email).first()
    else:
        target_user = db.query(User).filter(User.phone == request.phone).first()
    
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Check if this is a manual reset (no OTP provided) or OTP reset
    if not request.otp_code:
        # Manual password reset - check permissions
        if current_user.role == "superadmin":
            # Superadmin can reset any password
            pass
        elif current_user.role == "admin":
            # Admin can reset dispatcher, driver, user roles only
            if target_user.role in ["admin", "superadmin"]:
                raise HTTPException(status_code=403, detail="Admin cannot reset another admin or superadmin password")
        elif current_user.role == "dispatcher":
            # Dispatcher cannot manually reset anyone's password
            raise HTTPException(status_code=403, detail="Dispatcher cannot manually reset passwords")
        elif current_user.role in ["driver", "user"]:
            # Drivers and users can only reset their own password via OTP
            if target_user.id != current_user.id:
                raise HTTPException(status_code=403, detail="You can only reset your own password")
            raise HTTPException(status_code=400, detail="OTP required for password reset")
    else:
        # OTP reset - users can only reset their own password
        if target_user.id != current_user.id and current_user.role not in ["superadmin", "admin"]:
            raise HTTPException(status_code=403, detail="You can only reset your own password")
        
        # Find valid OTP (must not be used yet)
        otp = db.query(OTP).filter(
            OTP.user_id == target_user.id,
            OTP.otp_code == request.otp_code,
            OTP.is_used == False,
            OTP.expires_at > datetime.utcnow()
        ).first()
        
        if not otp:
            raise HTTPException(status_code=400, detail="Invalid or expired OTP")
        
        # Mark OTP as used
        otp.is_used = True
    
    # Update password
    target_user.hashed_password = get_password_hash(request.new_password)
    target_user.updated_at = datetime.utcnow()
    
    db.commit()
    
    return {"message": "Password reset successfully"}

@app.get("/auth/active-sessions", response_model=List[ActiveSessionWithUser])
async def get_active_sessions(current_user: User = Depends(require_admin_or_higher), db: Session = Depends(get_db)):
    """Get all active sessions (admin and superadmin only)"""
    sessions = db.query(ActiveSession).options(joinedload(ActiveSession.user)).all()
    
    result = []
    for session in sessions:
        session_with_user = ActiveSessionWithUser(
            id=session.id,
            user_id=session.user_id,
            login_at=session.login_at,
            logout_at=session.logout_at,
            ip_address=session.ip_address,
            is_active=session.is_active,
            user_name=session.user.name,
            user_email=session.user.email,
            user_role=session.user.role
        )
        result.append(session_with_user)
    
    return result

# User management endpoints
@app.get("/users/", response_model=List[UserSchema])
async def get_users(current_user: User = Depends(require_admin_or_higher), db: Session = Depends(get_db)):
    """Get all users (admin and superadmin only)"""
    # Dispatchers cannot access users endpoints
    if current_user.role == "dispatcher":
        raise HTTPException(status_code=403, detail="Dispatchers cannot access users endpoints")
    users = db.query(User).all()
    return users

@app.post("/users/", response_model=UserSchema)
async def create_user(user: UserCreate, current_user: User = Depends(require_admin_or_higher), db: Session = Depends(get_db)):
    """Create a new user (admin and superadmin only)"""
    # Dispatchers cannot access users endpoints
    if current_user.role == "dispatcher":
        raise HTTPException(status_code=403, detail="Dispatchers cannot access users endpoints")
    
    # Admins cannot create other admins or superadmins
    if current_user.role == "admin" and user.role in ["admin", "superadmin"]:
        raise HTTPException(status_code=403, detail="Admins cannot create admin or superadmin users")
    # Check if email already exists
    existing_user = db.query(User).filter(User.email == user.email).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Create user
    hashed_password = get_password_hash(user.password)
    db_user = User(
        name=user.name,
        email=user.email,
        phone=user.phone,
        hashed_password=hashed_password,
        role=user.role,
        is_active=user.is_active
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    
    return db_user

@app.put("/users/{user_id}", response_model=UserSchema)
async def update_user(user_id: int, user_update: UserUpdate, current_user: User = Depends(require_admin_or_higher), db: Session = Depends(get_db)):
    """Update a user (admin and superadmin only)"""
    # Dispatchers cannot access users endpoints
    if current_user.role == "dispatcher":
        raise HTTPException(status_code=403, detail="Dispatchers cannot access users endpoints")
    
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Admins cannot update other admins or superadmins
    if current_user.role == "admin" and user.role in ["admin", "superadmin"]:
        raise HTTPException(status_code=403, detail="Admins cannot update admin or superadmin users")
    
    # Admins cannot promote users to admin or superadmin
    if current_user.role == "admin" and user_update.role in ["admin", "superadmin"]:
        raise HTTPException(status_code=403, detail="Admins cannot promote users to admin or superadmin")
    
    # Update fields
    update_data = user_update.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(user, field, value)
    
    user.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(user)
    
    return user

@app.delete("/users/{user_id}")
async def delete_user(user_id: int, current_user: User = Depends(require_admin_or_higher), db: Session = Depends(get_db)):
    """Delete a user (admin and superadmin only)"""
    # Dispatchers cannot access users endpoints
    if current_user.role == "dispatcher":
        raise HTTPException(status_code=403, detail="Dispatchers cannot access users endpoints")
    
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Admins cannot delete other admins or superadmins
    if current_user.role == "admin" and user.role in ["admin", "superadmin"]:
        raise HTTPException(status_code=403, detail="Admins cannot delete admin or superadmin users")
    
    # Users cannot delete themselves
    if user.id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    db.delete(user)
    db.commit()
    
    return {"message": "User deleted successfully"}

@app.get("/users/me", response_model=UserProfile)
async def get_my_profile(current_user: User = Depends(get_current_user)):
    """Get own profile"""
    return UserProfile(
        id=current_user.id,
        name=current_user.name,
        email=current_user.email,
        phone=current_user.phone,
        role=current_user.role,
        is_active=current_user.is_active,
        last_login=current_user.last_login,
        created_at=current_user.created_at
    )

@app.put("/users/{user_id}/reset-password")
async def manual_password_reset(user_id: int, new_password: str = "", current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Manually reset a user's password (role-based rules apply)"""
    # Find target user
    target_user = db.query(User).filter(User.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Apply role-based rules
    if current_user.role == "superadmin":
        # Superadmin can reset any password
        pass
    elif current_user.role == "admin":
        # Admin can reset dispatcher, driver, user roles only
        if target_user.role in ["admin", "superadmin"]:
            raise HTTPException(status_code=403, detail="Admin cannot reset another admin or superadmin password")
    elif current_user.role == "dispatcher":
        # Dispatcher cannot manually reset anyone's password
        raise HTTPException(status_code=403, detail="Dispatcher cannot manually reset passwords")
    else:
        # Drivers and users cannot reset passwords
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    
    # Update password
    target_user.hashed_password = get_password_hash(new_password)
    target_user.updated_at = datetime.utcnow()
    
    db.commit()
    
    return {"message": "Password reset successfully"}

@app.put("/users/me/password")
async def change_my_password(password_change: PasswordChange, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Change own password"""
    if not verify_password(password_change.current_password, current_user.hashed_password):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    
    current_user.hashed_password = get_password_hash(password_change.new_password)
    current_user.updated_at = datetime.utcnow()
    db.commit()
    
    return {"message": "Password changed successfully"}

# User requests endpoints
@app.post("/requests/", response_model=UserRequestSchema)
async def create_user_request(request: UserRequestCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Create user request"""
    user_request = UserRequest(
        user_id=current_user.id,
        request_type=request.request_type,
        order_id=request.order_id,
        message=request.message,
        status="pending"
    )
    db.add(user_request)
    db.commit()
    db.refresh(user_request)
    
    return user_request

@app.get("/requests/", response_model=List[UserRequestSchema])
async def get_user_requests(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Get user requests"""
    if current_user.role in ["admin", "superadmin"]:
        # Admins and superadmins see all requests
        requests = db.query(UserRequest).all()
    else:
        # Regular users see only their requests
        requests = db.query(UserRequest).filter(UserRequest.user_id == current_user.id).all()
    
    return requests

@app.put("/requests/{request_id}", response_model=UserRequestSchema)
async def update_user_request(request_id: int, request_update: UserRequestUpdate, current_user: User = Depends(require_admin_or_higher), db: Session = Depends(get_db)):
    """Update user request (approve/reject) - admin and superadmin only"""
    user_request = db.query(UserRequest).filter(UserRequest.id == request_id).first()
    if not user_request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    user_request.status = request_update.status
    db.commit()
    db.refresh(user_request)
    
    return user_request

# Existing endpoints (with auth protection)
@app.get("/drivers", response_model=List[DriverSchema])
async def get_drivers_endpoint(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Get all drivers"""
    drivers = get_drivers(db)
    return drivers

@app.post("/drivers", response_model=DriverSchema)
async def create_driver_endpoint(driver: DriverCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Create a new driver"""
    return create_driver(db, driver)

@app.get("/vehicles", response_model=List[VehicleSchema])
async def get_vehicles_endpoint(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Get all vehicles"""
    vehicles = get_vehicles(db)
    return vehicles

@app.post("/vehicles", response_model=VehicleSchema)
async def create_vehicle_endpoint(vehicle: VehicleCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Create a new vehicle"""
    return create_vehicle(db, vehicle)

@app.get("/orders", response_model=List[OrderSchema])
async def get_orders_endpoint(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Get all orders"""
    orders = get_orders(db)
    return orders

@app.post("/orders", response_model=OrderSchema)
async def create_order_endpoint(order: OrderCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Create a new order"""
    return create_order(db, order)

@app.post("/optimize/", response_model=OptimizationResponse)
async def optimize_route_endpoint(request: OptimizationRequest, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Optimize route for given driver, vehicle, and orders"""
    try:
        # Get pending orders
        pending_orders = get_pending_orders(db)
        
        # Filter orders by the ones provided in the request
        selected_orders = [order for order in pending_orders if order.id in request.order_ids]
        
        if not selected_orders:
            return OptimizationResponse(
                success=False,
                message="No valid orders found for optimization",
                optimized_routes=[]
            )
        
        # Optimize routes
        optimized_routes = optimize_routes(request.driver_id, request.vehicle_id, selected_orders, db)
        
        # Clear existing routes for this driver and vehicle
        from models import Route
        db.query(Route).filter(
            Route.driver_id == request.driver_id,
            Route.vehicle_id == request.vehicle_id
        ).delete()
        
        # Save optimized routes to database
        saved_routes = []
        for i, route_data in enumerate(optimized_routes):
            route = Route(
                route_name=f"Route {request.driver_id}-{request.vehicle_id}",
                driver_id=request.driver_id,
                vehicle_id=request.vehicle_id,
                order_id=route_data["order_id"],
                sequence_number=i + 1,
                stop_type=route_data["stop_type"],
                estimated_arrival=route_data.get("estimated_arrival"),
                estimated_departure=route_data.get("estimated_departure"),
                distance_from_previous=route_data.get("distance_from_previous"),
                travel_time_from_previous=route_data.get("travel_time_from_previous"),
                status="planned"
            )
            db.add(route)
            saved_routes.append(route)
        
        db.commit()
        
        # Convert to response format
        response_routes = []
        for route in saved_routes:
            response_routes.append(RouteSchema.from_orm(route))
        
        return OptimizationResponse(
            success=True,
            message=f"Route optimized successfully with {len(saved_routes)} stops",
            optimized_routes=response_routes
        )
        
    except Exception as e:
        return OptimizationResponse(
            success=False,
            message=f"Optimization failed: {str(e)}",
            optimized_routes=[]
        )

@app.get("/reports/summary")
async def get_reports_summary(current_user: User = Depends(require_dispatcher_or_higher), db: Session = Depends(get_db)):
    """Get reports summary (superadmin, admin, dispatcher only)"""
    
    # Get today's date
    today = datetime.utcnow().date()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)
    
    # Mock data - replace with actual queries
    summary = {
        "dispatches": {
            "today": 0,
            "this_week": 0,
            "this_month": 0
        },
        "deliveries": {
            "delivered": 0,
            "yet_to_deliver": 0,
            "cancelled": 0,
            "reassigned": 0
        },
        "drivers": {
            "total_active_today": 0,
            "parcels_per_driver": [
                {"driver_id": 1, "driver_name": "John", "total_parcels": 5, "delivered": 3, "pending": 2}
            ]
        },
        "parcels": {
            "total_today": 0,
            "unassigned": 0
        }
    }
    
    return summary

# Reports endpoints
@app.get("/reports/export")
async def export_reports(db: Session = Depends(get_db)):
    """Generate and return XLSX report"""
    # Get summary data
    summary = await get_reports_summary(db)
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Sheet 1: Dispatch Summary
    ws_dispatch = wb.create_sheet("Dispatch Summary")
    ws_dispatch.append(["Period", "Total Dispatches"])
    ws_dispatch.append(["Today", summary["dispatches"]["today"]])
    ws_dispatch.append(["This Week", summary["dispatches"]["this_week"]])
    ws_dispatch.append(["This Month", summary["dispatches"]["this_month"]])
    
    # Make headers bold
    for cell in ws_dispatch[1]:
        cell.font = Font(bold=True)
    
    # Sheet 2: Delivery Status
    ws_delivery = wb.create_sheet("Delivery Status")
    ws_delivery.append(["Status", "Count"])
    ws_delivery.append(["Delivered", summary["deliveries"]["delivered"]])
    ws_delivery.append(["Yet to Deliver", summary["deliveries"]["yet_to_deliver"]])
    ws_delivery.append(["Cancelled", summary["deliveries"]["cancelled"]])
    ws_delivery.append(["Reassigned", summary["deliveries"]["reassigned"]])
    
    # Make headers bold
    for cell in ws_delivery[1]:
        cell.font = Font(bold=True)
    
    # Sheet 3: Driver Summary
    ws_drivers = wb.create_sheet("Driver Summary")
    ws_drivers.append(["Driver Name", "Total Parcels", "Delivered", "Pending", "Cancelled", "Completion %"])
    
    for driver in summary["drivers"]["parcels_per_driver"]:
        completion_pct = (driver["delivered"] / driver["total_parcels"] * 100) if driver["total_parcels"] > 0 else 0
        ws_drivers.append([
            driver["driver_name"],
            driver["total_parcels"],
            driver["delivered"],
            driver["pending"],
            0,  # Cancelled - not tracked in current data
            f"{completion_pct:.1f}%"
        ])
    
    # Make headers bold
    for cell in ws_drivers[1]:
        cell.font = Font(bold=True)
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return file
    headers = {
        "Content-Disposition": "attachment; filename=report.xlsx",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )

@app.post("/reports/upload")
async def upload_report(file: UploadFile = File(...)):
    """Upload XLSX report file"""
    # Validate file extension
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Only .xlsx files are allowed")
    
    # Create uploads directory if it doesn't exist
    upload_dir = "uploads/reports"
    os.makedirs(upload_dir, exist_ok=True)
    
    # Save file
    file_path = os.path.join(upload_dir, file.filename)
    
    try:
        with open(file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        return {"message": "Report uploaded successfully", "filename": file.filename}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to save file: {str(e)}")

if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8001, reload=True)
